import pandas as pd
from openpyxl import Workbook

def parse_data_to_workbook(actor_file, movie_file, output_file):
    # Read actor data into a pandas DataFrame
    actor_df = pd.read_csv(actor_file, sep='\t')
    
    # Read movie data into a pandas DataFrame
    movie_df = pd.read_csv(movie_file, sep='\t')
    
    # Create a new Excel workbook
    wb = Workbook()
    
    # Create a worksheet for actor data
    actor_ws = wb.create_sheet(title="Actor Data")
    for r_idx, row in enumerate(actor_df.iterrows(), start=1):
        for c_idx, value in enumerate(row[1], start=1):
            actor_ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Create a worksheet for movie data
    movie_ws = wb.create_sheet(title="Movie Data")
    for r_idx, row in enumerate(movie_df.iterrows(), start=1):
        for c_idx, value in enumerate(row[1], start=1):
            movie_ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Save the workbook to the specified output file
    wb.remove(wb.active)  # Remove the default sheet
    wb.save(output_file)
    print(f"Data saved to {output_file}")

# Example usage
parse_data_to_workbook("actors.tsv", "movies.tsv", "output.xlsx")
